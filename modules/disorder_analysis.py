import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
import streamlit as st
from utils.logger import logger
import random


all_disorders = [
            "depression", "trauma", "abuse", "anxiety", "stress", "psychosis",
            "personality disorder", "cognitive learning difficulties", "eating_disorder",
            "physical_problems", "addiction", "bereavement loss", "self esteem",
            "interpersonal_relationships", "living_welfare", "work_academic"
            ]

def disorder_ana(data):
    # Define the story_id and load the dataset
    # story_id = "story 8"
    # data = pd.read_csv("repdata.csv")

    # Create DataFrame
    df = pd.DataFrame(data)

    # Group by `disorder` and count the number of sentences
    disorder_counts = df.groupby('disorder').size().reset_index(name='sentence_count')

    # Calculate the total number of sentences
    total_sentences = len(df)

    # Add a column for the proportion of sentences
    disorder_counts['proportion'] = disorder_counts['sentence_count'] / total_sentences

    # Sort by sentence count in descending order
    disorder_counts = disorder_counts.sort_values(by='proportion', ascending=False, ignore_index=True)

    # Display the result
    # print("Analysis of Disorder Focus in the Story")
    # print(disorder_counts)
    
    # Ask the user for a custom proportion threshold or use the default value of 0.3
    # user_input = input(f"\nEnter a proportion threshold for disorder categorization (default is 0.3): ")
    # # If the user provides input, use it; otherwise, default to 0.3
    # try:
    #     proportion_threshold = float(user_input) if user_input else 0.3
    # except ValueError:
    #     print("Invalid input, using default threshold of 0.3.")
    #     proportion_threshold = 0.3
    proportion_threshold = 0.3
    # Identify disorders with proportion greater than the threshold
    focused_disorders = disorder_counts[disorder_counts['proportion'] > proportion_threshold]

    # If multiple disorders meet the condition, they will be categorized in the anthology
    if not focused_disorders.empty:
        focused_disorder_summary = f"""The story is categorized under the following disorders:"""
    else:
        focused_disorder_summary  = f"\nThe story is not disorder-focused."

    print("Analysis of Disorder Focus in the Story")
    # print(disorder_counts)
    return disorder_counts, focused_disorder_summary, focused_disorders

#Fill the empty rows with previous disorder names to avoid empty disorder rows
def propagate_disorder(df):
    last_valid_disorder = None

    # Iterate through each row and fill missing disorders
    for i in range(len(df)):
        if pd.notna(df.at[i, 'disorder']) and df.at[i, 'disorder'] != "":
            last_valid_disorder = df.at[i, 'disorder']
        elif last_valid_disorder is not None:
            df.at[i, 'disorder'] = last_valid_disorder
    
    return df

#creating story anthology
def create_anthology(disorder_counts, focused_disorders, story_id, anthology_file_path):
    
    """check if anthology already exists, if not create it"""
    
    # If the file already passes, read it; otherwise, create a new DataFrame
    if os.path.exists(anthology_file_path):
        anthology_df = pd.read_excel(anthology_file_path)
    else:
        # Initialize the DataFrame with all possible disorders and empty story_id columns
        anthology_df = pd.DataFrame({"disorder": all_disorders, "story_id": [None] * len(all_disorders), "proportion": [0.0] * len(all_disorders)})
        
    # print("before propagating throught disorder rows looking for null rows")
    # print(anthology_df)
    anthology_df = propagate_disorder(anthology_df)
    # print("after propagating and filling the null rows with previous valid disorder row")
    # print(anthology_df)
    # Iterate through the disorders identified in the current story
    for _, row in focused_disorders.iterrows():
        disorder = row['disorder']
        proportion = row['proportion']

        # Check if the same story_id exists under the current disorder
        existing_entry = anthology_df[(anthology_df['disorder'] == disorder) & (anthology_df['story_id'] == story_id)]
        
        # If the same story_id doesn't exist, check if any story_id exists under this disorder
        existing_any_story_id = anthology_df[(anthology_df['disorder'] == disorder) & (anthology_df['story_id'].notna())]
        if not existing_entry.empty:
            # Update only the proportion for the specific disorder and story_id
            anthology_df.loc[(anthology_df['disorder'] == disorder) & (anthology_df['story_id'] == story_id), 'proportion'] = proportion
        elif not existing_any_story_id.empty:
            # Create a new row for the disorder and story_id
            # Add a new row for this disorder and story_id
            new_row = {"disorder": disorder, "story_id": story_id, "proportion": proportion}
            anthology_df = pd.concat([anthology_df, pd.DataFrame([new_row])], ignore_index=True)
        else:
            anthology_df.loc[anthology_df['disorder']==disorder, 'story_id'] = story_id
            anthology_df.loc[anthology_df['disorder'] == disorder, 'proportion'] = proportion

    print(anthology_df)
    print(f"\nThe story '{story_id}' has been added or updated in the anthology under the following disorders:")
    print(disorder_counts[['disorder', 'proportion']].to_string(index=False))
    # Now, use OpenPyxl to merge `disorder` cells and separate `story_id` rows
    wb = Workbook()
    ws = wb.active

    # Write headers
    ws.append(["disorder", "story_id", "proportion"])

    # Group the anthology DataFrame by `disorder` and list all associated `story_id`s
    grouped = anthology_df.groupby("disorder")

    start_row = 2
    for disorder, group in grouped:
        # Sort the group by proportion in descending order for each disorder
        group = group.sort_values(by='proportion', ascending=False)
        story_ids = group['story_id'].tolist()
        proportion = group['proportion'].tolist()  # Get the correct proportion for the current disorder

        if story_ids:
            # Write the disorder name in the first column
            ws[f"A{start_row}"] = disorder
            # Write each story_id in the second column (new row for each story_id)
            for i, story_id in enumerate(story_ids):
                ws[f"B{start_row}"] = story_id
                # Ensure proportion is a single scalar value for each story_id
                # proportion_to_write = proportion[i] if proportion[i] is not None else 1
                ws[f"C{start_row}"] = proportion[i]
                start_row += 1
            # Merge the disorder cells in column A for the range of story_ids
            ws.merge_cells(start_row=start_row - len(story_ids), start_column=1, end_row=start_row - 1, end_column=1)

    # Apply alignment to the merged cells
    for row in range(2, start_row):
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top")
        ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="top")

    # Save the workbook
    output_file = "story_anthology.xlsx"
    print(output_file)
    wb.save(output_file)
    # print(f"\nThe story anthology has been updated and saved to '{output_file}'.")
    return output_file


#Finding the story which ranks first for the given disorder
def story_for_disorder_from_anthology(anthology_data, disorder):
    try:
        # Define the path to the anthology CSV file
        # anthology_file = "story_anthology.csv"

        # Load the dataset
        anth_data = pd.read_excel(anthology_data)
        # print("anthology data")
        # print(anth_data)
        anth_data =propagate_disorder(anth_data)
        # print("after propagating")
        # print(anth_data)
        # Group by `disorder` and `story_id`, and calculate the proportion for each story
        disorder_counts1 = anth_data.groupby(['disorder', 'story_id']).agg({'proportion': 'max'}).reset_index()

        # Ask the user to input a disorder
        # user_disorder = input("Enter a disorder to fetch the story with the highest proportion: ").strip()
        user_disorder = disorder.strip()
        # Check if the disorder exists in the dataset
        if user_disorder in disorder_counts1['disorder'].values:
            # Filter the data for the selected disorder
            selected_disorder_data = disorder_counts1[disorder_counts1['disorder'] == user_disorder]

            # Get the story with the highest proportion for that disorder
            top_story = selected_disorder_data.loc[selected_disorder_data['proportion'].idxmax()]
            return top_story['story_id'], top_story['proportion']
        else:
            return None, None
    except Exception as e:
        print(f"An error occurred while processing the data: {str(e)}")
        return None, None
        
            #     print(f"\nThe story with the highest proportion under '{user_disorder}' is:")
            #     print(f"Story ID: {top_story['story_id']}")
            #     print(f"Proportion: {top_story['proportion']}")
            # else:
            #     print(f"\nThe disorder '{user_disorder}' is not found in the dataset.")


def generate_plot(disorder_counts, story_id):
    # Plotting a horizontal bar chart
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.barh(disorder_counts['disorder'], disorder_counts['sentence_count'], color='skyblue')
    ax.set_xlabel('Sentence Count')
    ax.set_ylabel('Disorder')
    ax.set_title(f'Sentence Count by Disorder for story, {story_id}')
    ax.invert_yaxis()  # Invert y-axis to display the highest count at the top
    return fig

#Disorder design ui
def disorder_design():
    #Disorder_analysis
    # st.subheader("Analysis of Disorder Focus in the Story")
    st.write("")
    st.write("")
    st.write("")
    st.markdown('<h3 class="stTitle">Analysis of Disorder Focus in the Story</h3>', unsafe_allow_html=True)
    st.write("")
    st.write("")
    disorder_file = st.file_uploader("Upload a disorder analysis excel file", type=["csv"])
    if disorder_file is not None:
        st.session_state["disorder_file"] = disorder_file
    st.write("")
    anthology_file = st.file_uploader("Upload Existing Anthology XLSX (Optional)", type = ['xlsx'], accept_multiple_files = False)
    if anthology_file is not None:
        st.session_state["anthology_file"] = anthology_file
    story_id = st.text_input("Enter story ID for updating anthology: ")
    st.write("")
    disorderanalysis = st.button("Start Disorder Analysis & Update Anthology")
    st.write("")
    # Update_anthology = st.button("Update Anthology")
    if disorderanalysis:
        if disorder_file is not None:
            try:
                #need to provide story_id as well
                disorder_data = pd.read_csv(disorder_file)
                st.session_state["disorder_file"] = disorder_file
                st.session_state["disorder_data"] = disorder_data
                disorder_counts, focused_disorder_summary, focused_disorders = disorder_ana(disorder_data)
                st.session_state["disorder_counts"] = disorder_counts
                st.session_state["focused_disorder_summary"] = focused_disorder_summary
                st.session_state["focused_disorders"] = focused_disorders
                 # Logic to get an existing story_id or generate a random one
                story_id = story_id if story_id else f"story_{random.randint(1, 100)}"
                disorder_fig  = generate_plot(disorder_counts, story_id)
                st.session_state["disorder_fig"] = disorder_fig
               
                # If story ID is provided, create or update the anthology
                # anthology_file_path = anthology_file.name if anthology_file else "story_anthology.xlsx"
                # Option to upload an existing anthology CSV file (optional)
                if anthology_file is not None:
                    # If the user uploads a custom anthology file, use that
                    # anthology_file_path = os.path.join("uploaded_anthology.xlsx")
                    anthology_file_path = "uploaded_anthology.xlsx"
                    with open(anthology_file_path, "wb") as f:
                        f.write(anthology_file.getbuffer())
                else:
                    # Default anthology file name
                    anthology_file_path = "story_anthology.xlsx"
                st.session_state.anthology_file_path = anthology_file_path
                st.session_state.anthology_file = anthology_file

                output_file = create_anthology(disorder_counts, focused_disorders, story_id, anthology_file_path)
                
                # st.session_state["output_file"] = output_file
                #Display the results
                st.pyplot(disorder_fig)
                st.write(disorder_counts)
                st.write(focused_disorder_summary)
                st.dataframe(focused_disorders)
                st.download_button(
                                label="Download Anthology File",
                                data=open(output_file, "rb").read(),
                                file_name=output_file,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                st.success(f"The story anthology has been updated and saved to '{output_file}'.")
            except Exception as e:
                st.error(f"An error occurred while processing the disorder file")
                logger.error(f"Error details: {e}", exc_info=True)
    else:
        # If a file was processed earlier, display the saved results
        if st.session_state.get("disorder_counts") is not None:
            disorder_counts = st.session_state["disorder_counts"]
            focused_disorder_summary = st.session_state["focused_disorder_summary"]
            focused_disorders = st.session_state["focused_disorders"]
            disorder_fig = st.session_state["disorder_fig"]
            # st.subheader("Analysis of Disorder Focus in the Story")
            st.markdown('<h3 class="stTitle">Analysis of Disorder Focus in the Story</h3>', unsafe_allow_html=True)
            st.write("")
            st.write("")
            st.pyplot(disorder_fig)
            st.write(disorder_counts)
            st.write(focused_disorder_summary)
            st.dataframe(focused_disorders)
        else:
            st.info("Upload a file to start the analysis.")


    ####Disorder Focus Identification######
    # st.subheader("Disorder-Focused Story Identification")
    st.write("")
    st.write("")
    st.write("")
    st.markdown('<h3 class="stTitle">Disorder-Focused Story Identification</h3>', unsafe_allow_html=True)
    st.write("")
    st.write("")
    # all_disorders = [
    #         "depression", "trauma", "abuse", "anxiety", "stress", "psychosis",
    #         "personality disorder", "cognitive learning difficulties", "eating_disorder",
    #         "physical_problems", "addiction", "bereavement loss", "self esteem",
    #         "interpersonal_relationships", "living_welfare", "work_academic"
    #         ]
    # Input for the disorder
    disorder = st.selectbox("Select a disorder to fetch the story with the highest proportion:", ["Select a disorder"] + all_disorders)
    st.session_state["selected_disorder"] = disorder
    st.write("")
    disorderfocusedstory = st.button("Find the Top-Ranked story for the specified disorder")
    st.write("")
    if disorderfocusedstory:
        if st.session_state["anthology_file"] is not None:
            # Validate the uploaded anthology file
            try:
                anthology_data = pd.read_excel(st.session_state["anthology_file"], engine='openpyxl')
                st.session_state["anthology_data"] = anthology_data
                #check for required columns
                required_columns = {"disorder", "story_id", "proportion"}
                st.session_state["required_columns"] = required_columns
                anthology_data.columns = anthology_data.columns.str.strip().str.lower()
                st.session_state["anthology_data.columns"] = anthology_data.columns
                if not required_columns.issubset(set(anthology_data.columns)):
                    st.error("The anthology file must have 'disorder','story_id', and 'proportion' columns.")

                elif disorder != "Select a disorder":
                    # Fetch the story with the highest proportion for the entered disorder
                    story_id_focus, proportion = story_for_disorder_from_anthology(anthology_file, disorder)
                    st.session_state["story_id_focus"] = story_id_focus
                    st.session_state["proportion"] = proportion
                    st.session_state["disorder"] = disorder
                    if story_id_focus and proportion:
                        st.success(f"The story with the highest proportion under '{disorder}' is:")
                        st.write(f"**Story ID:** {story_id_focus}")
                        st.write(f"**Proportion:** {proportion}")
                    else:
                        st.warning(f"The disorder '{disorder}' is not found in the dataset.")
                else:
                    st.error("Please select a disorder.")
            except Exception as e:
                st.error(f"An error occurred while processing the file: {str(e)}")
        else:
            st.info("Please upload the anthology excel file.")
    else:
        if st.session_state.get("story_id_focus") is not None:
            story_id_focus = st.session_state["story_id_focus"]
            proportion = st.session_state["proportion"]
            st.session_state["disorder"] = disorder
            st.success(f"The story with the highest proportion under '{disorder}' is:")
            st.write(f"**Story ID:** {story_id_focus}")
            st.write(f"**Proportion:** {proportion}")

        st.info("Upload the anthology excel file to start the analysis.")
